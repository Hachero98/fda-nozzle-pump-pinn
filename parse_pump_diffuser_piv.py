"""
parse_pump_diffuser_piv.py
==========================
Parse the FDA Blood Pump diffuser PIV xlsx files (C1, C2, C4, C5, C6) into
flat CSVs suitable for the diffuser PINN.

The xlsx files share a common layout:
  - row 1, columns B..:  X coordinates (m)
  - column A, rows 2..:  Y coordinates (m)
  - cells (i,j):         velocity magnitude |V| (m/s); None outside fluid

For files with both Mean and st_dev sheets (C2, C4, C5), we capture both.

Output (one CSV per condition):
  x, y, v_mag, v_mag_std
where v_mag_std is empty if not provided.

Frame: stationary lab frame (housing-aligned), 2D plane through diffuser.
"""

from pathlib import Path
import csv
import openpyxl

ROOT = Path(__file__).parent
SRC  = ROOT / "Reference Papers/FDA_Dataset/FDA_OSEL_Benchmarks/Blood Pump/Data/Diffuser"
OUT  = ROOT / "Reference Papers/FDA_Dataset/Pump"
OUT.mkdir(parents=True, exist_ok=True)

CASES = {
    "C1": ("Mean_velocity_diffuser_C1.xlsx",            None),
    "C2": ("MeanSTD_velocity_diffuser_C2_2022.xlsx",    "MeanSTD"),
    "C4": ("MeanSTD_velocity_diffuser_C4_2022.xlsx",    "MeanSTD"),
    "C5": ("MeanSTD_velocity_diffuser_C5_2022.xlsx",    "MeanSTD"),
    "C6": ("Mean_velocity_diffuser_C6.xlsx",            None),
}


def _read_grid(ws):
    """Read a single 'X across, Y down' grid sheet -> (x_vals, y_vals, V[Y,X])."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], []
    header = rows[0]
    # X coords: header columns 1..N (skip A1 label)
    x_vals = [c for c in header[1:] if isinstance(c, (int, float))]
    y_vals, grid = [], []
    for r in rows[1:]:
        y = r[0]
        if not isinstance(y, (int, float)):
            continue
        y_vals.append(y)
        grid.append(list(r[1:1 + len(x_vals)]))
    return x_vals, y_vals, grid


def _sheet_named(wb, candidates):
    """Pick the first sheet whose name is in candidates, else the first sheet."""
    for c in candidates:
        if c in wb.sheetnames:
            return wb[c]
    return wb[wb.sheetnames[0]]


def parse_condition(case, fname, kind):
    src = SRC / fname
    if not src.exists():
        print(f"  SKIP {case}: {src.name} not found")
        return 0
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    if kind == "MeanSTD":
        mean_ws = _sheet_named(wb, ["Mean", "Sheet1"])
        std_ws  = _sheet_named(wb, ["st_dev", "St_Dev", "Sheet2"])
        x_m, y_m, V_m = _read_grid(mean_ws)
        x_s, y_s, V_s = _read_grid(std_ws)
        # Best-effort alignment: assume identical X/Y grids
        use_std = (x_m == x_s and y_m == y_s)
    else:
        ws = _sheet_named(wb, ["Sheet1", "Mean"])
        x_m, y_m, V_m = _read_grid(ws)
        V_s = None
        use_std = False

    wb.close()

    # Physical sanity bounds for |V| (m/s). Tip speed at 3500 RPM is 9.5 m/s;
    # diffuser flow is bulk-decelerating, so anything above ~12 m/s or below 0
    # is a PIV cross-correlation failure.
    V_MIN, V_MAX = 0.0, 15.0

    out_csv = OUT / f"PumpDiffuser_{case}.csv"
    n_valid, n_dropped = 0, 0
    with open(out_csv, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["x_m", "y_m", "v_mag_m_per_s", "v_mag_std_m_per_s"])
        for i, y in enumerate(y_m):
            for j, x in enumerate(x_m):
                v = V_m[i][j]
                if v is None:
                    continue
                if not isinstance(v, (int, float)):
                    continue
                if not (V_MIN <= v <= V_MAX):
                    n_dropped += 1
                    continue
                if use_std:
                    s = V_s[i][j] if (i < len(V_s) and j < len(V_s[i])) else None
                    s_str = f"{s:.6g}" if isinstance(s, (int, float)) else ""
                else:
                    s_str = ""
                w.writerow([f"{x:.6g}", f"{y:.6g}", f"{v:.6g}", s_str])
                n_valid += 1

    xmin, xmax = min(x_m), max(x_m)
    ymin, ymax = min(y_m), max(y_m)
    vmax_obs = max(
        v for row in V_m for v in row if isinstance(v, (int, float))
    )
    print(f"  {case}: {n_valid:5d} pts  ({n_dropped} outliers dropped)  "
          f"x=[{xmin:.4f},{xmax:.4f}]m  y=[{ymin:.4f},{ymax:.4f}]m  "
          f"-> {out_csv.name}")
    return n_valid


def main():
    print(f"Source: {SRC}")
    print(f"Output: {OUT}\n")
    total = 0
    for case, (fname, kind) in CASES.items():
        total += parse_condition(case, fname, kind)
    print(f"\nTotal: {total} valid PIV points across {len(CASES)} conditions.")


if __name__ == "__main__":
    main()
